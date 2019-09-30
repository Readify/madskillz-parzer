#!/usr/bin/env python3
import sys
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Fill
from openpyxl.worksheet.datavalidation import DataValidation
from urllib.request import urlopen
import re

def import_file_from_readify_github():
    logging.info('Getting latest Consulting.md...')
    response = urlopen('https://raw.githubusercontent.com/Readify/madskillz/master/Development.md')
    logging.info('Getting latest Consulting.md...Done!')
    return response

def change_cell_value(row, column, ws, value):
    ws.cell(row = row, column = column).value = value

def change_cell_style_header(row, column, ws):
    ws.cell(row = row, column = column).font = Font(size = 14, bold = True)

def change_cell_style_role_description(row, column, ws):
    ws.cell(row = row, column = column).font = Font(size = 13)

def get_role():
    roles = ("DataAndAnalytics", "Development", "Engineering", "InnovationAndDesign", "Managed Services", "ProgramManager")
    return input("Enter Role " + str(roles) + ": ")

def main():
    role = get_role()
    print (role)
    filename = 'MadSkillz.xlsx'
    role_pattern = '^## .+'
    role_description_pattern = '>.+'
    behaviour_pattern = '^- .+'
    row_number = 0
    wb = Workbook()
    active_ws = wb.active
    data_validation = DataValidation(type="list", formula1='"Cannot Assess,Need Coaching,Maturing,No Brainer,Outstanding"', allow_blank=False)
    f = import_file_from_readify_github()
    logging.info('Parsing...')
    for line in f:
        line = line.decode('unicode-escape')
        row_number += 1
        role = re.search(role_pattern, line)
        if role:
            active_ws = wb.create_sheet(0)
            row_number = 1
            active_ws.add_data_validation(data_validation)
            active_ws.title = line.strip('#')
            line = line.strip('#')
            change_cell_value(row_number, 1, active_ws, line)
            change_cell_style_header(row_number, 1, active_ws)
            change_cell_style_header(row_number, 2, active_ws)
            change_cell_value(row_number, 2, active_ws, "Rating")


        behaviour = re.search(behaviour_pattern, line)
        if behaviour:
            change_cell_value(row_number, 1, active_ws, line)
            rating_cell = active_ws.cell(row = row_number, column = 2)
            data_validation.add(rating_cell)

        role_description = re.search(role_description_pattern, line)
        if role_description:
            line = line.strip('>')
            change_cell_value(row_number, 1, active_ws, line)
            change_cell_style_role_description(row_number, 1, active_ws)

        c = active_ws.cell(row = row_number, column = 1)
        c.value = line

    logging.info('Parsing...Done!')
    wb.save(filename)
    logging.info('Saved "%s"', filename)

if __name__ == "__main__":
    logging.basicConfig(stream=sys.stdout, level=logging.INFO)
    main()
